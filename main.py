import openpyxl
import pymongo

client = pymongo.MongoClient("mongodb://127.0.0.1:27017/?compressors=disabled&gssapiServiceName=mongodb")
db = client['tgdb']



excel_file = openpyxl.load_workbook('2.xlsx')
ws = excel_file.active

name = ws['C']
stock = ws['E']
price = ws['F']
for i in range(2,len(name)):
    print(name[i].value,stock[i].value,price[i].value)
    db.tgdb.insert_one({'id':i,'name':name[i].value, 'stock':stock[i].value, 'price':price[i].value})

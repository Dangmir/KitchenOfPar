import telebot
from telebot import types
import pymongo
from pprint import pprint
import re
import openpyxl


client = pymongo.MongoClient("mongodb://127.0.0.1:27017/?compressors=disabled&gssapiServiceName=mongodb")
dbuser = client['tguser']
db = client['tgdb']


bot = telebot.TeleBot('1840665419:AAH_pw1tKDJt_LkP_-mO2wOMTsVDKAsr4Bg',parse_mode=None)

#Регистраиция
@bot.message_handler(commands=['start'])
def start(message):
    a = dbuser.user.find_one({'tgid': message.from_user.id})
    print(a,message.from_user.id)
    if a:
        bot.send_message(message.from_user.id,'Вы уже зарегистрированны')
        shop(message)
    else:
        a = bot.send_message(message.from_user.id,'Здравствуйте, введите ваши данные перед заказом через запятую вот в таком формате:\nAlexey,89877500711')
        bot.register_next_step_handler(a,register)

def register(message):
    text = message.text.split(',')
    name = text[0]
    phone = text[1]
    dbuser.user.insert_one({'name':name,'phone':phone,'tgid':message.from_user.id})
    bot.send_message(message.from_user.id,'Успешно зарегистрированны')



@bot.message_handler(commands=['shop'])
def shop(message):
    mainmenu = types.ReplyKeyboardMarkup(resize_keyboard=True,row_width=2)
    key1 = types.KeyboardButton(text='Одноразки')
    key2 = types.KeyboardButton(text='Жидкости')
    key3 = types.KeyboardButton(text='Поиск')
    mainmenu.add(key1, key2,key3)
    bot.send_message(message.from_user.id, 'Это главное меню!', reply_markup=mainmenu)



@bot.callback_query_handler(func=lambda c: True)
def ans(c):
    if c.data=='toBucket':
        toBucket(c)
    if c.data =='buyall':
        buyall(c)
    if c.data == 'delete_from_bucket':
        delete_from_bucket(c)
    if c.data == 'right':
        right(c)
    if c.data == 'left':
        left(c)
#МЕНЮ
def menuone(message):
    mainmenu = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key1 = types.KeyboardButton(text='HQD')
    key2 = types.KeyboardButton(text='IZI')
    key3 = types.KeyboardButton(text='Корзина')
    key4 = types.KeyboardButton(text='Назад')
    mainmenu.add(key1, key2,key3,key4)
    bot.send_message(message.from_user.id, 'Одноразки', reply_markup=mainmenu)

def hqdsearch(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD'}}).limit(10)
    mainmenu = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key1 = types.KeyboardButton(text='HQD Cuvie 300 затяжек')
    key2 = types.KeyboardButton(text='HQD Super 600 затяжек')
    key3 = types.KeyboardButton(text='HQD Cuvie Plus 1200 затяжек')
    key4 = types.KeyboardButton(text='HQD King 2000 затяжек')
    key5 = types.KeyboardButton(text='HQD Maxx 2500 затяжек')
    key6 = types.KeyboardButton(text='Корзина')
    key7 = types.KeyboardButton(text='Вернуться')
    key8 = types.KeyboardButton(text='Назад')
    mainmenu.add(key1, key2, key3,key4,key5,key6,key7,key8)
    bot.send_message(message.from_user.id, 'HQD', reply_markup=mainmenu)

def hqd300(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD Cuvie 280'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1 )
        bot.send_message(message.from_user.id,f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',reply_markup=mainmenu)

def hqd600(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD Super'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def hqd1200(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD Cuvie PLUS'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def hqd2000(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD King'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def hqd2500(message):
    result = db.tgdb.find({"name": {"$regex": 'HQD Maxx'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def izisearch(message):
    mainmenu = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key1 = types.KeyboardButton(text='IZI X3 1200 затяжек')
    key2 = types.KeyboardButton(text='IZI X8 1500 затяжек')
    key3 = types.KeyboardButton(text='IZI MAX 1600 затяжек')
    key6 = types.KeyboardButton(text='Корзина')
    key7 = types.KeyboardButton(text='Вернуться')
    key8 = types.KeyboardButton(text='Назад')
    mainmenu.add(key1, key2, key3,key6, key7, key8)
    bot.send_message(message.from_user.id, 'IZI', reply_markup=mainmenu)

def izi1200(message):
    result = db.tgdb.find({"name": {"$regex": 'IZI X3'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def izi1500(message):
    result = db.tgdb.find({"name": {"$regex": 'IZI X8'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)

def izi1600(message):
    result = db.tgdb.find({"name": {"$regex": 'IZI MAX'}}).limit(10)
    for i in result:
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
        mainmenu.add(key1)
        bot.send_message(message.from_user.id,
                         f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]} \nОстаток:{i["stock"]}',
                         reply_markup=mainmenu)


def delete_from_bucket(message):
    text = message.message.text
    id = text.split()[0].split(':')[1]
    item = dbuser.user.find_one({'tgid': message.from_user.id})
    itemas = item['bucket']
    d = []

    dbuser.user.update_one({'tgid':message.from_user.id},{'$pull':{'bucket':{'id':int(id)}}})
    for i in itemas:
        d.append(int(i['price']))
    suma = sum(d)
    bot.delete_message(message.message.chat.id,message.message.id)

    a = bot.send_message(message.from_user.id,f'Сумма заказа: {suma}')

def backet(message):
    mainmenu = types.InlineKeyboardMarkup()
    delete = types.InlineKeyboardButton(text='Удалить',callback_data='delete_from_bucket')
    mainmenu.add(delete)
    buy = types.InlineKeyboardMarkup()
    buy1 = types.InlineKeyboardButton(text='Оформить заказ',callback_data='buyall')
    buy.add(buy1)
    try:
        a = dbuser.user.find_one({'tgid':message.from_user.id})['bucket']
        print(len(a))
        if len(a)!= 0:
            print(len(a))
            bot.send_message(message.from_user.id, 'Корзина')
            count = 1
            item = []
            for i in a:
                item.append(int(i['price']))
                print(i)

                bot.send_message(message.from_user.id,
                                 f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]}\nОстаток на складе: {i["stock"]}\n',
                                 reply_markup=mainmenu)
            sumzak = sum(item)
            print(sumzak)
            bot.send_message(message.from_user.id, f'Заказ\nСумма заказа: {sumzak}', reply_markup=buy)
        else:
            print("KILLL")
            bot.send_message(message.from_user.id,'Корзина пуста')
    except BaseException as e:
        print(e)
        bot.send_message(message.from_user.id, "Корзина пуста")


def toBucket(message):
    id_tovara = message.message.text.split()[0].split(':')[1]
    tovar = db.tgdb.find_one({'id':int(id_tovara)})
    user = dbuser.user.find_one({'tgid':message.from_user.id})
    print(tovar)
    item = {'name':tovar['name'],'price':tovar['price'],'id':tovar['id'],'stock':tovar['stock']}
    dbuser.user.update_one({'tgid':message.from_user.id},{'$push':{'bucket':item}})
    bot.send_message(message.from_user.id,'Добавленно в корзину')

def buyall(message):
    id = dbuser.user.find_one({'tgid':message.from_user.id})
    bucket = id['bucket']
    print(bucket)
    bot.send_message(861813649, f'Hello Pidar прими заказ от {id["name"]}\n Номер телефона:{id["phone"]}')
    item = []
    for i in bucket:
        item.append(int(i['price']))
        bot.send_message(861813649,f'{i["name"]}')
    sumzak = sum(item)
    bot.send_message(861813649,f'Сумма заказа: {sumzak}\n-----------------------------')
    dbuser.user.update_one({'tgid':message.from_user.id},{'$unset':{'bucket':1}})
    bot.send_message(message.from_user.id,'Заказ оформлен ждите звонка')




def search(message):
    a = bot.send_message(message.from_user.id,"Введите ваш запрос")
    bot.register_next_step_handler(a,watch_result_search,offset=0,page=0)

def watch_result_search(message,offset,page):
    db.tgdb.create_index([('name', pymongo.TEXT)], name='search_index')
    result = db.tgdb.find({'$text': {'$search': message.text}})
    result_limit = db.tgdb.find({'$text': {'$search': message.text}}).skip(offset).limit(5)
    mainmenu = types.InlineKeyboardMarkup()
    key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
    mainmenu.add(key1)
    count_page = 0
    for i in result:
        count_page+=1
    for j in result_limit:
        bot.send_message(message.from_user.id,f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]}\nОстаток на складе: {i["stock"]}\n',reply_markup=mainmenu)
    count_page = round(count_page/5)
    mainmenu = types.InlineKeyboardMarkup()
    key1 = types.InlineKeyboardButton(text='>',callback_data=f'right')
    key2 = types.InlineKeyboardButton(text='<',callback_data=f'left')
    mainmenu.add(key2,key1)
    bot.send_message(message.from_user.id,f'Запрос:{message.text}\nКоличество страниц: {count_page}\nТекущая страница:{page}',reply_markup=mainmenu)
def right(message):
    name = message.message.text.split(':')[1].split('\n')[0]
    offset = int(message.message.text.split(':')[3])
    mainmenu = types.InlineKeyboardMarkup()
    key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
    mainmenu.add(key1)

    contpage = int(message.message.text.split(':')[2].split('\n')[0].strip())
    if offset > contpage:
        bot.send_message(message.from_user.id,'Конец поиска')
    else:
        print(contpage)
        result_limit = db.tgdb.find({'$text': {'$search': name}}).skip(offset*5).limit(5)

        for i in result_limit:
            bot.send_message(message.from_user.id,f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]}\nОстаток на складе: {i["stock"]}\n',reply_markup=mainmenu)
        offset+=1
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text='>', callback_data=f'right')
        key2 = types.InlineKeyboardButton(text='<', callback_data=f'left')
        mainmenu.add(key2, key1)
        bot.send_message(message.from_user.id,
                         f'Запрос:{name}\nКоличество страниц: {contpage}\nТекущая страница:{offset}',
                         reply_markup=mainmenu)

def left(message):
    mainmenu = types.InlineKeyboardMarkup()
    key1 = types.InlineKeyboardButton(text=f'В корзину', callback_data='toBucket')
    mainmenu.add(key1)
    name = message.message.text.split(':')[1].split('\n')[0]
    offset = int(message.message.text.split(':')[3])

    contpage = int(message.message.text.split(':')[2].split('\n')[0].strip())
    if offset > contpage:
        bot.send_message(message.from_user.id,'Конец поиска')
    else:
        print(contpage)
        result_limit = db.tgdb.find({'$text': {'$search': name}}).skip((offset-1)*5).limit(5)

        for i in result_limit:
            bot.send_message(message.from_user.id,f'ID:{i["id"]}\nНазвание:{i["name"]}\nЦена: {i["price"]}\nОстаток на складе: {i["stock"]}\n',reply_markup=mainmenu)
        offset-=1
        mainmenu = types.InlineKeyboardMarkup()
        key1 = types.InlineKeyboardButton(text='>', callback_data=f'right')
        key2 = types.InlineKeyboardButton(text='<', callback_data=f'left')
        mainmenu.add(key2, key1)
        bot.send_message(message.from_user.id,
                         f'Запрос:{name}\nКоличество страниц: {contpage}\nТекущая страница:{offset}',
                         reply_markup=mainmenu)


@bot.message_handler(content_types=['document'])
def document(message):
    print(message.from_user.id)
    if message.from_user.id == 628229833:
        f = 0
        try:
            fileinfo = bot.get_file(message.document.file_id)
            download_file = bot.download_file(fileinfo.file_path)
            with open('3.xlsx','wb') as new_file:
                new_file.write(download_file)
            excel_file = openpyxl.load_workbook('3.xlsx')
            ws = excel_file.active

            name = ws['C']
            stock = ws['E']
            price = ws['F']

            for i in range(2, len(name)):
                a = db.tgdb.find_one({'name':name[i].value})
                try:
                    if a['stock'] != stock[i].value:
                        print(stock[i].value)
                        db.tgdb.update({'name':name[i].value},{'id':a['id'],'name':name[i].value,'stock':stock[i].value,'price':price[i].value})
                        f+=1
                        bot.send_message(message.from_user.id,f'Заменено позиций:{f},Заменено:{name[i].value}')
                except:
                    pass
        except BaseException as e:
            print(e)
    else:
        bot.send_message(message.from_user.id,'Вам это не доступно')




@bot.message_handler(content_types=['text'])
def handler(message):
    if message.text == 'Поиск':
        search(message)
    if message.text == 'Одноразки':
        menuone(message)
    if message.text == 'HQD':
        hqdsearch(message)
    if message.text == 'Корзина':
        backet(message)
    if message.text == 'HQD Cuvie 300 затяжек':
        hqd300(message)
    if message.text == 'HQD Super 600 затяжек':
        hqd600(message)
    if message.text == 'HQD Cuvie Plus 1200 затяжек':
        hqd1200(message)
    if message.text == 'HQD King 2000 затяжек':
        hqd2000(message)
    if message.text == 'HQD Maxx 2500 затяжек':
        hqd2500(message)
    if message.text == "IZI X3 1200 затяжек":
        izi1200(message)
    if message.text == "IZI X8 1500 затяжек":
        izi1500(message)
    if message.text == "IZI MAX 1600 затяжек":
        izi1600(message)
    if message.text =='Назад':
        shop(message)
    if message.text == 'Вернуться':
        menuone(message)
    if message.text == "IZI":
        izisearch(message)

if __name__ == '__main__':
    while True:
        try:
            bot.polling()
        except BaseException as e:
            print(e)
            pass
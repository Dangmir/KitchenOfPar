from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import os
import pymongo
from xls2xlsx import XLS2XLSX
import os
from selenium.webdriver.chrome.options import Options
options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
import openpyxl

client = pymongo.MongoClient("mongodb://127.0.0.1:27017/?compressors=disabled&gssapiServiceName=mongodb")
tgdb = client['tgdb']

EXE_PATH = r'/usr/bin/chromedriver'
driver = webdriver.Chrome(executable_path=EXE_PATH)
def get_stocks():

    driver.get('https://online.moysklad.ru/')
    name = driver.find_element_by_name('j_username').send_keys('kirito@profi-di')
    password = driver.find_element_by_name('j_password').send_keys('ce385701.')
    submit = driver.find_element_by_name('submitButton').click()
    time.sleep(4)
    a = driver.find_element_by_xpath('//*[@id="site"]/table/tbody/tr[3]/td/div/table/tbody/tr/td[1]/div/table/tbody/tr/td[6]/table/tbody/tr[1]/td/img').click()
    time.sleep(1)
    b = driver.find_element_by_xpath('//*[@id="site"]/table/tbody/tr[3]/td/div/div[2]/div/span[2]/a/span').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="site"]/table/tbody/tr[4]/td/table/tbody/tr/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[1]/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr/td[3]/div/table/tbody/tr/td[1]/img').click()
    time.sleep(1)
    driver.find_elements_by_class_name('gwt-MenuItem')[4].click()
    path = 'C:/Users/lexa324sa/Downloads'
    time.sleep(7)
    a = []
    for root, dirs, files in os.walk(path):
        a.append(files)
    return a

def update_db(a):
    print(a[-1])
    name = a[-1]
    print(name)
    try:
        x2x = XLS2XLSX('C:/Users/lexa324sa/Downloads/'+name)
        x2x.to_xlsx('C:/Users/lexa324sa/Downloads/'+name+'x')
        excel_file = openpyxl.load_workbook('C:/Users/lexa324sa/Downloads/'+name+'x')
        ws = excel_file.active

        name = ws['C']
        stock = ws['E']
        price = ws['F']
        tgdb.tgdb.remove()
        for i in range(2,len(name)):
            id = int(i)
            stocks = stock[i].value
            prices = price[i].value
            names = name[i].value
            tgdb.tgdb.insert_one({'id':id,'stock':stocks,'price':prices,'name':names})
    except BaseException as e:
        print(e)

update_db(get_stocks()[-1])
driver.close()